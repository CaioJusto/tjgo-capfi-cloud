from __future__ import annotations

import re
from io import BytesIO

from openpyxl import load_workbook


PROCESS_PATTERN = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}")


def extract_process_numbers_from_xlsx(contents: bytes) -> list[str]:
    workbook = load_workbook(filename=BytesIO(contents), read_only=True, data_only=True)
    worksheet = workbook.active
    found: list[str] = []
    seen: set[str] = set()

    for row in worksheet.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            for match in PROCESS_PATTERN.findall(text):
                if match not in seen:
                    seen.add(match)
                    found.append(match)

    return found
